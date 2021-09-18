from openpyxl import Workbook
from openpyxl import load_workbook
import random as rd
import sys
from wx import MessageBox

__all__ = ["Data"]

class Data(object):

#   学生名单用.xlsx格式存的
    

    def __init__(self):
        self.openFileFlg = False
        self.openFile("./学生名单.xlsx")


    def openFile(self, dirname_):
        try:
            self.dirname = dirname_
            self.wb = load_workbook(self.dirname)
            self.ws = self.wb.active
        except Exception:
            MessageBox("打开文件失败")
        else:
            self.openFileFlg = True

        try:
            self.nameTuple = self.getNamesByRow_1()
#            print(self.nameTuple)
            if self.nameTuple == 0:
                raise Exception("")
        except Exception:
            MessageBox("\"姓名\"一栏格式不正确")
        else:
            print(self.getNamesByRow_1())
            self.stuNum = len(self.nameTuple)
            self.vstList = [0 for i in range(0, self.stuNum + 1)]
            self.vstNum = 0
            self.shuffledListGen = self.createGen()


    def getNamesByRow_1(self):
        it = iter(self.ws["1"])
        nameCell = self.ws["A1"]
        cnt = 0
        for cell in it:
            if (cell.value in ("名字", "姓名")):
                cnt += 1
                if cnt == 0:
                    nameCell = cell
        if cnt == 0:
            MessageBox("文件中没有姓名一栏，请检查文件")
            return 0
        elif cnt > 1:
            MessageBox("您是想整爷是吧？给爷爪巴")
            return 0
        else:
            col = self.ws[chr(nameCell.column + 65 - 1)]
            length = len(col)
            return col[1:length + 1]
                

#   正常模式
    def getRandName_Normal(self):
        num = rd.randint(1, self.stuNum)
        nameString = self.nameTuple[num - 1].value
        return nameString


#   无重复模式
    def createGen(self):
        shuffledList = list(self.nameTuple).copy()
        rd.shuffle(shuffledList)
        for i in shuffledList:
            yield i

    def getRandName_NoRepetition(self):
        try:
            nameString = next(self.shuffledListGen).value
        except StopIteration:
            self.shuffledListGen = self.createGen()
            nameString = next(self.shuffledListGen).value
        return nameString
